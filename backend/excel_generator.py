import io
import re
import tempfile
import zipfile
from pathlib import Path

import openpyxl

from models import CONCEPTOS_ADMIN, CONCEPTOS_LUMISIAL, TesoreríaData

TEMPLATE_PATH = Path(__file__).parent / "template" / "FORMATO_TESORERIA_2026.xlsm"

_PITCH_FAMILY_RE = re.compile(rb'pitchFamily="(\d+)"')


def _fix_pitch_family(value: bytes) -> bytes:
    """Clamp pitchFamily values to openpyxl's max of 52."""
    def clamp(m: re.Match) -> bytes:
        v = int(m.group(1))
        return b'pitchFamily="' + str(min(v, 52)).encode() + b'"'
    return _PITCH_FAMILY_RE.sub(clamp, value)


def _load_template() -> openpyxl.Workbook:
    """Load the template, patching invalid pitchFamily values in drawings."""
    with zipfile.ZipFile(TEMPLATE_PATH) as zin:
        names = zin.namelist()
        drawing_names = {n for n in names if "drawings/drawing" in n and n.endswith(".xml")}

        with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                data = zin.read(name)
                if name in drawing_names:
                    data = _fix_pitch_family(data)
                zout.writestr(name, data)

    wb = openpyxl.load_workbook(tmp_path, keep_vba=True)
    tmp_path.unlink(missing_ok=True)
    return wb

# Row where the member table starts
MEMBER_START_ROW = 19
# Row where each member's block starts in 1006-Ingresos (3 rows per member)
INGRESOS_BLOCK_START = 11
INGRESOS_BLOCK_SIZE = 3


def generate_excel(data: TesoreríaData) -> io.BytesIO:
    wb = _load_template()

    _fill_ingreso_datos(wb["INGRESO DE DATOS"], data)
    _fill_1006_ingresos(wb["1006-Ingresos"], data)
    _fill_1007_egresos(wb["1007 - Egresos"], data)
    _fill_1004_cuadre(wb["1004 - Cuadre de caja"], data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fill_ingreso_datos(ws, data: TesoreríaData):
    org = data.organizacion

    ws["E10"] = org.diocesis
    ws["I8"] = org.santuario
    ws["I9"] = org.codigo_santuario
    # openpyxl expects a Python datetime/date for date-formatted cells
    ws["I10"] = org.mes_anio

    # Members
    for i, m in enumerate(data.miembros):
        row = MEMBER_START_ROW + i
        ws.cell(row=row, column=3).value = m.nombre.upper()  # C
        ws.cell(row=row, column=4).value = m.cedula           # D
        ws.cell(row=row, column=5).value = m.aportes          # E
        if m.num_cuotas > 1:
            ws.cell(row=row, column=6).value = m.num_cuotas   # F

    # Otras entradas: OBOLO always at row 22; free items 23+
    ws["I22"] = next(
        (e.valor for e in data.otras_entradas if e.concepto.upper() == "OBOLO"), 0
    )
    free_otras = [e for e in data.otras_entradas if e.concepto.upper() != "OBOLO"]
    for i, entrada in enumerate(free_otras):
        row = 23 + i
        ws.cell(row=row, column=8).value = entrada.concepto  # H
        ws.cell(row=row, column=9).value = entrada.valor     # I

    # Salidas de administración (rows 22–28, cols K=11, M=13)
    admin = data.salidas_admin
    for i in range(7):
        row = 22 + i
        if i < len(admin) and admin[i].valor:
            if admin[i].fecha:
                ws.cell(row=row, column=11).value = admin[i].fecha  # K
            ws.cell(row=row, column=13).value = admin[i].valor      # M

    # Salidas 1008 (rows 30–46, L=12, M=13); row 31 = OBOLO
    obolo_1008 = next(
        (s.valor for s in data.salidas_1008 if s.concepto.upper() == "OBOLO"), 0
    )
    ws["M31"] = obolo_1008
    free_1008 = [s for s in data.salidas_1008 if s.concepto.upper() != "OBOLO"]
    for i, salida in enumerate(free_1008):
        row = 32 + i
        ws.cell(row=row, column=12).value = salida.concepto  # L
        ws.cell(row=row, column=13).value = salida.valor     # M

    # Salidas lumisial (rows 51–75, L=12, M=13)
    # Pre-filled concepts at rows 51–58; free rows start at 59
    lumisial_map = {c.upper(): c for c in CONCEPTOS_LUMISIAL}
    pre_filled = []
    free_lumisial = []
    for s in data.salidas_lumisial:
        if s.concepto.upper() in lumisial_map:
            pre_filled.append(s)
        else:
            free_lumisial.append(s)

    for s in pre_filled:
        idx = CONCEPTOS_LUMISIAL.index(lumisial_map[s.concepto.upper()])
        row = 51 + idx
        ws.cell(row=row, column=13).value = s.valor  # M (concept already in template)

    for i, s in enumerate(free_lumisial):
        row = 59 + i
        ws.cell(row=row, column=12).value = s.concepto  # L
        ws.cell(row=row, column=13).value = s.valor     # M


def _fill_1006_ingresos(ws, data: TesoreríaData):
    for i, m in enumerate(data.miembros):
        block_row = INGRESOS_BLOCK_START + i * INGRESOS_BLOCK_SIZE

        ws.cell(row=block_row, column=1).value = m.fecha_recibo   # A — FECHA

        # Column D is the visible receipt number cell (B and C are hidden columns).
        # Cell D is integer-formatted — write only the consecutive part.
        num_parts = (m.num_recibo or "").split("-")
        try:
            ws.cell(row=block_row, column=4).value = int(num_parts[-1])
        except (ValueError, IndexError):
            pass

        for j, cv in enumerate(m.cuentas_varias[:3]):
            r = block_row + j
            ws.cell(row=r, column=11).value = cv.detalle  # K
            ws.cell(row=r, column=12).value = cv.valor    # L

    # Actividades Lumisial — row 461
    for i, act in enumerate(data.actividades_lumisial):
        row = 461 + i
        ws.cell(row=row, column=11).value = act.concepto  # K
        ws.cell(row=row, column=12).value = act.valor     # L


def _fill_1007_egresos(ws, data: TesoreríaData):
    # 1007-Egresos is fully formula-driven (pulls from INGRESO DE DATOS).
    # No manual cell writes needed.
    pass


def _fill_1004_cuadre(ws, data: TesoreríaData):
    cc = data.cuadre_caja

    # Receipt range  (row 10: "Recibos de caja del No. [B10] al No. [D10]")
    ws["B10"] = cc.recibo_desde
    ws["D10"] = cc.recibo_hasta

    # Saldo anterior (row 12: "SALDO ANTERIOR: [C12]"; used in E89 = +C12+C34-D87)
    ws["C12"] = cc.saldo_anterior

    # Voucher range  (row 37: "Comprobantes de pago del No. [B37] AL No. [D37]")
    ws["B37"] = cc.voucher_desde
    ws["D37"] = cc.voucher_hasta
